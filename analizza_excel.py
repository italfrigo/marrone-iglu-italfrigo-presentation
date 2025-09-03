import pandas as pd
import os
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart

# Percorso del file Excel originale
file_originale = 'rh_conti.xls'

# Funzione per leggere il file Excel e stampare informazioni sulle sue schede
def analizza_file_excel(file_path):
    print(f"Analisi del file: {file_path}")
    try:
        # Apri il file Excel con xlrd per vedere i nomi dei fogli
        workbook = xlrd.open_workbook(file_path)
        sheet_names = workbook.sheet_names()
        
        print(f"Il file contiene {len(sheet_names)} fogli:")
        for i, name in enumerate(sheet_names):
            sheet = workbook.sheet_by_name(name)
            print(f"  {i+1}. {name} - Righe: {sheet.nrows}, Colonne: {sheet.ncols}")
            
            # Mostra un'anteprima delle prime righe di ciascun foglio
            if sheet.nrows > 0:
                print(f"    Anteprima delle prime righe:")
                for row_idx in range(min(5, sheet.nrows)):
                    row_values = []
                    for col_idx in range(min(5, sheet.ncols)):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row_values.append(str(cell_value))
                    print(f"    Riga {row_idx+1}: {', '.join(row_values)}")
                print()
        
        return sheet_names
    except Exception as e:
        print(f"Errore durante l'analisi del file: {e}")
        return []

# Funzione per creare una versione migliorata del file Excel
def crea_excel_migliorato(file_originale, sheet_names):
    print("\nCreazione della versione migliorata del file Excel...")
    
    # Nome del nuovo file
    nuovo_file = 'RH_Corso_Venezia_Offerta.xlsx'
    
    # Crea un nuovo workbook con openpyxl
    wb = Workbook()
    
    # Rimuovi il foglio predefinito
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Leggi i dati dal file originale e crea fogli migliorati
    for sheet_name in sheet_names:
        try:
            # Leggi i dati con pandas
            df = pd.read_excel(file_originale, sheet_name=sheet_name)
            
            # Crea un nuovo foglio nel workbook
            ws = wb.create_sheet(title=sheet_name)
            
            # Aggiungi intestazioni
            for col_idx, column in enumerate(df.columns):
                cell = ws.cell(row=1, column=col_idx+1)
                cell.value = column
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            
            # Aggiungi i dati
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row):
                    cell = ws.cell(row=row_idx+2, column=col_idx+1)
                    cell.value = value
                    
                    # Allineamento e bordi
                    cell.alignment = Alignment(horizontal='left')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Stile alternato per le righe
                    if row_idx % 2 == 1:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Adatta la larghezza delle colonne
            for col_idx in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 15
                
            print(f"Foglio '{sheet_name}' creato con successo.")
            
            # Se ci sono dati numerici, aggiungi un grafico
            numeric_columns = df.select_dtypes(include=['number']).columns
            if len(numeric_columns) > 0 and len(df) > 0:
                # Crea un grafico a barre per la prima colonna numerica
                chart = BarChart()
                chart.title = f"Grafico {sheet_name}"
                chart.y_axis.title = numeric_columns[0]
                chart.x_axis.title = df.columns[0] if len(df.columns) > 0 else "Elementi"
                
                data = Reference(ws, min_col=numeric_columns.get_loc(numeric_columns[0])+1, 
                                min_row=1, max_row=len(df)+1, max_col=numeric_columns.get_loc(numeric_columns[0])+1)
                cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                # Posiziona il grafico
                ws.add_chart(chart, "H2")
                print(f"Grafico aggiunto al foglio '{sheet_name}'")
                
        except Exception as e:
            print(f"Errore durante la creazione del foglio '{sheet_name}': {e}")
    
    # Salva il nuovo file
    wb.save(nuovo_file)
    print(f"\nFile migliorato salvato come: {nuovo_file}")
    return nuovo_file

# Esegui l'analisi
sheet_names = analizza_file_excel(file_originale)

# Crea la versione migliorata
if sheet_names:
    nuovo_file = crea_excel_migliorato(file_originale, sheet_names)
    print(f"\nPer visualizzare il file migliorato, puoi aprirlo con un programma come LibreOffice Calc o Microsoft Excel.")
else:
    print("Non è stato possibile analizzare il file Excel.")
